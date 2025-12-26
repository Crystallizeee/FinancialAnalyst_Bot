"""
Portfolio Analysis and Export Module
Provides diversification analysis and PDF/Excel export
"""
import json
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import sys
sys.path.append('..')
from config.settings import DATA_DIR


# Set matplotlib to use non-interactive backend
plt.switch_backend('Agg')


class PortfolioAnalyzer:
    """Analyze portfolio diversification and generate reports"""
    
    def __init__(self):
        pass
    
    def analyze_diversification(self, holdings: Dict[str, Any], stock_info: Dict[str, Any]) -> Dict[str, Any]:
        """
        Analyze portfolio diversification
        
        Args:
            holdings: User's holdings {stock: {quantity, avg_price, ...}}
            stock_info: Additional stock info {stock: {sector, market_cap, ...}}
            
        Returns:
            Diversification analysis results
        """
        if not holdings:
            return {"error": "Portfolio kosong"}
        
        # Calculate total value
        total_value = 0
        stock_values = {}
        sectors = {}
        
        for stock, data in holdings.items():
            qty = data.get("quantity", 0)
            price = stock_info.get(stock, {}).get("current_price", data.get("avg_price", 0))
            value = qty * price
            total_value += value
            stock_values[stock] = value
            
            # Get sector
            sector = stock_info.get(stock, {}).get("sector", "Unknown")
            if sector not in sectors:
                sectors[sector] = 0
            sectors[sector] += value
        
        if total_value == 0:
            return {"error": "Tidak ada nilai portfolio"}
        
        # Calculate allocations
        stock_allocation = {stock: (value / total_value * 100) for stock, value in stock_values.items()}
        sector_allocation = {sector: (value / total_value * 100) for sector, value in sectors.items()}
        
        # Diversification score (0-100)
        # Based on: number of stocks, sector spread, concentration
        num_stocks = len(holdings)
        num_sectors = len(sectors)
        
        # Concentration: Herfindahl index (lower = more diversified)
        hhi = sum((alloc/100)**2 for alloc in stock_allocation.values())
        
        # Scores
        stock_score = min(num_stocks * 10, 30)  # Max 30 for 3+ stocks
        sector_score = min(num_sectors * 15, 30)  # Max 30 for 2+ sectors
        concentration_score = max(0, 40 - (hhi * 100))  # Max 40 for low concentration
        
        diversification_score = stock_score + sector_score + concentration_score
        
        # Rating
        if diversification_score >= 80:
            rating = "Sangat Terdiversifikasi ✅"
            recommendation = "Portfolio kamu sudah terdiversifikasi dengan baik!"
        elif diversification_score >= 60:
            rating = "Terdiversifikasi 👍"
            recommendation = "Portfolio cukup terdiversifikasi. Pertimbangkan menambah sektor lain."
        elif diversification_score >= 40:
            rating = "Perlu Diversifikasi ⚠️"
            recommendation = "Tambahkan lebih banyak saham dari sektor berbeda."
        else:
            rating = "Sangat Terkonsentrasi 🔴"
            recommendation = "Portfolio sangat terkonsentrasi! Segera diversifikasi untuk mengurangi risiko."
        
        # Top holding warning
        max_allocation = max(stock_allocation.values()) if stock_allocation else 0
        top_holding = max(stock_allocation.items(), key=lambda x: x[1])[0] if stock_allocation else None
        
        return {
            "total_value": total_value,
            "num_stocks": num_stocks,
            "num_sectors": num_sectors,
            "stock_allocation": stock_allocation,
            "sector_allocation": sector_allocation,
            "diversification_score": round(diversification_score, 1),
            "rating": rating,
            "recommendation": recommendation,
            "top_holding": top_holding,
            "top_holding_pct": max_allocation,
            "hhi": round(hhi, 4)
        }
    
    def format_diversification_analysis(self, analysis: Dict[str, Any]) -> str:
        """Format diversification analysis into readable summary"""
        if "error" in analysis:
            return f"❌ {analysis['error']}"
        
        lines = ["📊 ANALISIS DIVERSIFIKASI", "━━━━━━━━━━━━━━━━━━━━━━", ""]
        
        # Score
        score = analysis["diversification_score"]
        if score >= 80:
            score_bar = "🟢🟢🟢🟢🟢"
        elif score >= 60:
            score_bar = "🟢🟢🟢🟢⚪"
        elif score >= 40:
            score_bar = "🟢🟢🟢⚪⚪"
        else:
            score_bar = "🟢🟢⚪⚪⚪"
        
        lines.append(f"📈 Skor Diversifikasi: {score}/100")
        lines.append(f"   {score_bar}")
        lines.append(f"   {analysis['rating']}")
        lines.append("")
        
        # Summary
        lines.append(f"💼 Total Value: Rp {analysis['total_value']:,.0f}")
        lines.append(f"📦 Jumlah Saham: {analysis['num_stocks']}")
        lines.append(f"🏭 Jumlah Sektor: {analysis['num_sectors']}")
        lines.append("")
        
        # Top holding warning
        if analysis["top_holding_pct"] > 50:
            lines.append(f"⚠️ PERINGATAN: {analysis['top_holding']} = {analysis['top_holding_pct']:.1f}% portfolio")
            lines.append("   Terlalu terkonsentrasi pada satu saham!")
            lines.append("")
        
        # Sector allocation
        lines.append("📊 Alokasi per Sektor:")
        for sector, pct in sorted(analysis["sector_allocation"].items(), key=lambda x: x[1], reverse=True):
            bar_len = int(pct / 10)
            bar = "█" * bar_len
            lines.append(f"   • {sector}: {pct:.1f}% {bar}")
        lines.append("")
        
        # Stock allocation
        lines.append("📊 Alokasi per Saham:")
        for stock, pct in sorted(analysis["stock_allocation"].items(), key=lambda x: x[1], reverse=True)[:5]:
            lines.append(f"   • {stock}: {pct:.1f}%")
        lines.append("")
        
        # Recommendation
        lines.append("💡 Rekomendasi:")
        lines.append(f"   {analysis['recommendation']}")
        
        return "\n".join(lines)
    
    def generate_allocation_chart(self, analysis: Dict[str, Any]) -> Optional[BytesIO]:
        """Generate pie chart for portfolio allocation"""
        if "error" in analysis or not analysis.get("sector_allocation"):
            return None
        
        try:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6), facecolor='#1a1a2e')
            
            # Sector pie chart
            sectors = list(analysis["sector_allocation"].keys())
            sector_values = list(analysis["sector_allocation"].values())
            colors1 = plt.cm.Set3(range(len(sectors)))
            
            ax1.pie(sector_values, labels=sectors, autopct='%1.1f%%', colors=colors1,
                   textprops={'color': 'white'}, wedgeprops={'edgecolor': 'white', 'linewidth': 1})
            ax1.set_title('Alokasi per Sektor', color='white', fontsize=14, fontweight='bold')
            ax1.set_facecolor('#1a1a2e')
            
            # Stock pie chart
            stocks = list(analysis["stock_allocation"].keys())
            stock_values = list(analysis["stock_allocation"].values())
            colors2 = plt.cm.Set2(range(len(stocks)))
            
            ax2.pie(stock_values, labels=stocks, autopct='%1.1f%%', colors=colors2,
                   textprops={'color': 'white'}, wedgeprops={'edgecolor': 'white', 'linewidth': 1})
            ax2.set_title('Alokasi per Saham', color='white', fontsize=14, fontweight='bold')
            ax2.set_facecolor('#1a1a2e')
            
            plt.tight_layout()
            
            buf = BytesIO()
            fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                       facecolor='#1a1a2e', edgecolor='none')
            buf.seek(0)
            plt.close(fig)
            
            return buf
            
        except Exception as e:
            print(f"Error generating allocation chart: {e}")
            return None
    
    def export_to_excel(self, holdings: Dict[str, Any], transactions: List[Dict], 
                        stock_info: Dict[str, Any]) -> Optional[BytesIO]:
        """Export portfolio to Excel file"""
        try:
            wb = Workbook()
            
            # Holdings sheet
            ws_holdings = wb.active
            ws_holdings.title = "Holdings"
            
            # Headers
            headers = ["Saham", "Jumlah", "Avg Price", "Current Price", "Value", "P/L", "P/L %"]
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws_holdings.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            row = 2
            total_invested = 0
            total_current = 0
            
            for stock, data in holdings.items():
                qty = data.get("quantity", 0)
                avg_price = data.get("avg_price", 0)
                current_price = stock_info.get(stock, {}).get("current_price", avg_price)
                invested = qty * avg_price
                current_value = qty * current_price
                pnl = current_value - invested
                pnl_pct = (pnl / invested * 100) if invested else 0
                
                total_invested += invested
                total_current += current_value
                
                ws_holdings.cell(row=row, column=1, value=stock)
                ws_holdings.cell(row=row, column=2, value=qty)
                ws_holdings.cell(row=row, column=3, value=avg_price)
                ws_holdings.cell(row=row, column=4, value=current_price)
                ws_holdings.cell(row=row, column=5, value=current_value)
                ws_holdings.cell(row=row, column=6, value=pnl)
                ws_holdings.cell(row=row, column=7, value=f"{pnl_pct:.2f}%")
                
                row += 1
            
            # Total row
            ws_holdings.cell(row=row, column=1, value="TOTAL")
            ws_holdings.cell(row=row, column=5, value=total_current)
            ws_holdings.cell(row=row, column=6, value=total_current - total_invested)
            
            # Transactions sheet
            ws_tx = wb.create_sheet("Transactions")
            tx_headers = ["Date", "Type", "Stock", "Quantity", "Price", "Total", "P/L"]
            
            for col, header in enumerate(tx_headers, 1):
                cell = ws_tx.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row, tx in enumerate(transactions, 2):
                ws_tx.cell(row=row, column=1, value=tx.get("timestamp", "")[:10])
                ws_tx.cell(row=row, column=2, value=tx.get("type", ""))
                ws_tx.cell(row=row, column=3, value=tx.get("stock", ""))
                ws_tx.cell(row=row, column=4, value=tx.get("quantity", 0))
                ws_tx.cell(row=row, column=5, value=tx.get("price", 0))
                ws_tx.cell(row=row, column=6, value=tx.get("total", 0))
                ws_tx.cell(row=row, column=7, value=tx.get("profit", ""))
            
            # Auto-width columns
            for ws in [ws_holdings, ws_tx]:
                for column in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = max_length + 2
            
            # Save to BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            return buf
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None
    
    def export_to_pdf(self, holdings: Dict[str, Any], transactions: List[Dict],
                      stock_info: Dict[str, Any], analysis: Dict[str, Any]) -> Optional[BytesIO]:
        """Export portfolio to PDF file"""
        try:
            buf = BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=30)
            elements.append(Paragraph("Portfolio Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Summary
            if analysis and "error" not in analysis:
                elements.append(Paragraph("Diversification Score", styles['Heading2']))
                elements.append(Paragraph(f"Score: {analysis['diversification_score']}/100 - {analysis['rating']}", styles['Normal']))
                elements.append(Paragraph(f"Total Value: Rp {analysis['total_value']:,.0f}", styles['Normal']))
                elements.append(Spacer(1, 20))
            
            # Holdings table
            elements.append(Paragraph("Holdings", styles['Heading2']))
            
            table_data = [["Stock", "Qty", "Avg Price", "Current", "Value", "P/L %"]]
            
            for stock, data in holdings.items():
                qty = data.get("quantity", 0)
                avg_price = data.get("avg_price", 0)
                current_price = stock_info.get(stock, {}).get("current_price", avg_price)
                current_value = qty * current_price
                invested = qty * avg_price
                pnl_pct = ((current_value - invested) / invested * 100) if invested else 0
                
                table_data.append([
                    stock,
                    str(qty),
                    f"Rp {avg_price:,.0f}",
                    f"Rp {current_price:,.0f}",
                    f"Rp {current_value:,.0f}",
                    f"{pnl_pct:+.2f}%"
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            buf.seek(0)
            
            return buf
            
        except Exception as e:
            print(f"Error exporting to PDF: {e}")
            return None


# Singleton instance
portfolio_analyzer = PortfolioAnalyzer()
